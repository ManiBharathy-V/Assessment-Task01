from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from flask_cors import CORS
import os

app = Flask(__name__)
CORS(app)

# Define the Excel file path
EXCEL_FILE = 'user_data.xlsx'

# Function to initialize the workbook and sheet if they do not exist
def init_workbook():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        # Create header row
        headers = ['Employee Name', 'Employee Email', "Phone", 'Age', 'Designation', 'Address']
        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
        wb.save(EXCEL_FILE)

@app.route('/submit-form', methods=['POST'])
def submit_form():
    data = request.json
    user_data = data.get('user', {})

    name = user_data.get('employee name')
    email = user_data.get('employee email')
    phone = user_data.get('phone number')
    age = user_data.get('age')
    designation = user_data.get('designation')
    address = user_data.get('address')

    # Load the workbook and get the active sheet
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find the next empty row
    next_row = ws.max_row + 1

    # Append the form data to the sheet
    ws.append([name, email, phone, age, designation, address])

    # Save the workbook
    wb.save(EXCEL_FILE)

    return jsonify({'message': 'Form data saved successfully'}), 200

if __name__ == '__main__':
    init_workbook()
    app.run(debug=True)
